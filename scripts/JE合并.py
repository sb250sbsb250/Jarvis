import os
import pandas as pd
import re
from pathlib import Path
import xlrd
# 目标文件夹
folder = r"C:\Users\xinzh\Downloads\RObot\2026_06_05_北京擎朗科技有限公司-2023年12期月报"

# 收集所有 Excel 文件
files = list(Path(folder).glob("*.xls*"))
print(f"找到 {len(files)} 个文件")

all_data = []

# JE表标准表头
JE_HEADERS = [
    "日期", "会计年度", "期间", "凭证字", "凭证号", "摘要",
    "科目编码", "科目全名", "币别", "原币金额", "借方金额", "贷方金额",
    "制单", "审核", "过账", "出纳", "附件数", "来源系统", "业务类型",
    "审核状态", "作废状态"
]


def extract_company_name(filename):
    """
    从文件名提取公司名
    匹配模式：
    - "北京擎朗科技有限公司" 这样的公司名
    - 支持 _公司名 或 -公司名 格式
    """
    # 去掉扩展名
    name = Path(filename).stem

    # 模式1: 匹配 _公司名 或 -公司名 之后的部分
    # 例如: "2024年12期_北京擎朗科技有限公司_JE" -> "北京擎朗科技有限公司"
    match = re.search(r'[_-]([\u4e00-\u9fa5]+(?:有限|股份|集团|科技|实业|投资|控股)?公司[\u4e00-\u9fa5]*)', name)
    if match:
        return match.group(1)

    # 模式2: 匹配 _公司名 格式，公司名可能包含括号
    # 例如: "2024年12期_北京擎朗科技有限公司" -> "北京擎朗科技有限公司"
    match = re.search(
        r'[_-]([\u4e00-\u9fa5()（）]+(?:有限|股份|集团|科技|实业|投资|控股)?(?:责任)?公司[\u4e00-\u9fa5()（）]*)', name)
    if match:
        return match.group(1)

    # 模式3: 更宽松的匹配，取 _ 或 - 后包含"公司"的部分
    match = re.search(r'[_-]([^_-]*公司[^_-]*)', name)
    if match:
        return match.group(1)

    # 模式4: 整个文件名中匹配包含"公司"的部分
    match = re.search(r'([\u4e00-\u9fa5]+公司[\u4e00-\u9fa5]*)', name)
    if match:
        return match.group(1)

    # 如果都匹配不到，返回文件名
    return name


def find_je_header_row(df_raw):
    """查找JE表头行，返回行索引，找不到返回None"""
    for i in range(min(20, len(df_raw))):
        row = df_raw.iloc[i].astype(str).str.strip().tolist()
        # 检查是否包含"日期"和"凭证字"（JE表特征）
        if "日期" in row and "凭证字" in row:
            return i
        # 或者只检查"凭证字"
        if "凭证字" in row and "摘要" in row:
            return i
    return None


# 处理每个文件
for f in files:
    print(f"处理: {f.name}")
    try:
        # 提取公司名
        company_name = extract_company_name(f.name)
        print(f"  公司名: {company_name}")

        is_xls = f.suffix.lower() == '.xls'

        # 读取整个文件，找JE sheet（如果存在）
        try:
            # 先检查有哪些sheet
            if is_xls:
                wb = xlrd.open_workbook(f)
                sheet_names = wb.sheet_names()
                wb.release_resources()
            else:
                from openpyxl import load_workbook

                wb = load_workbook(f, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
        except:
            # 如果读取失败，尝试默认sheet
            sheet_names = [0]  # 默认第一个sheet

        # 查找JE相关的sheet
        je_sheet_name = None
        for sn in sheet_names:
            if isinstance(sn, str) and "JE" in sn.upper():
                je_sheet_name = sn
                break

        # 如果没找到JE sheet，尝试第一个sheet
        if je_sheet_name is None:
            if isinstance(sheet_names[0], str):
                je_sheet_name = sheet_names[0]
            else:
                je_sheet_name = 0  # 默认第一个sheet

        print(f"  使用sheet: {je_sheet_name}")

        # 读取sheet
        df_raw = pd.read_excel(
            f,
            sheet_name=je_sheet_name,
            header=None,
            dtype=object,
            engine='xlrd' if is_xls else 'openpyxl'
        )

        # 查找表头行
        header_row = find_je_header_row(df_raw)

        if header_row is None:
            print(f"  ⚠️ 未找到JE表头行，跳过")
            continue

        print(f"  表头行: 第{header_row + 1}行")

        # 重新读取，使用表头行
        df = pd.read_excel(
            f,
            sheet_name=je_sheet_name,
            header=header_row,
            dtype=object,
            engine='xlrd' if is_xls else 'openpyxl'
        )

        # 删除全空行
        df = df.dropna(how='all')

        if df.empty:
            print(f"  ⚠️ 数据为空，跳过")
            continue

        # 只保留JE标准表头中存在的列
        available_cols = [col for col in JE_HEADERS if col in df.columns]
        if len(available_cols) < 3:
            print(f"  ⚠️ 列匹配过少（{len(available_cols)}列），跳过")
            continue

        df = df[available_cols]

        # 清理：过滤掉表头重复行和无效行
        if "凭证字" in df.columns:
            df = df[df["凭证字"] != "凭证字"]

        # 过滤全空行（再次）
        df = df.dropna(how='all')

        # 过滤没有凭证号的行
        if "凭证号" in df.columns:
            df = df[df["凭证号"].notna()]
            df = df[df["凭证号"].astype(str).str.strip() != ""]

        # 添加公司名列
        df.insert(0, "公司名", company_name)

        # 添加文件名列
        df.insert(1, "源文件", f.name)

        all_data.append(df)
        print(f"  ✅ 读取 {len(df)} 行凭证数据")

    except Exception as e:
        print(f"  ❌ 错误: {e}")
        import traceback

        traceback.print_exc()

# 合并所有数据
if all_data:
    merged = pd.concat(all_data, ignore_index=True)

    # 调整列顺序
    column_order = [
        "公司名", "源文件", "日期", "会计年度", "期间", "凭证字", "凭证号",
        "摘要", "科目编码", "科目全名", "币别", "原币金额", "借方金额", "贷方金额",
        "制单", "审核", "过账", "出纳", "附件数", "来源系统", "业务类型",
        "审核状态", "作废状态"
    ]
    existing_columns = [col for col in column_order if col in merged.columns]
    merged = merged[existing_columns]

    # 转换日期列
    if "日期" in merged.columns:
        merged["日期"] = pd.to_datetime(merged["日期"], errors='coerce')

    # 转换金额列
    amount_cols = ["原币金额", "借方金额", "贷方金额"]
    for col in amount_cols:
        if col in merged.columns:
            merged[col] = pd.to_numeric(merged[col], errors='coerce').fillna(0)

    # 转换数值列
    if "凭证号" in merged.columns:
        merged["凭证号"] = pd.to_numeric(merged["凭证号"], errors='coerce')
    if "附件数" in merged.columns:
        merged["附件数"] = pd.to_numeric(merged["附件数"], errors='coerce')

    output_path = os.path.join(folder, "JE合并结果.xlsx")
    merged.to_excel(output_path, index=False)

    print(f"\n{'=' * 50}")
    print(f"✅ 完成！")
    print(f"   文件数: {len(all_data)}")
    print(f"   总行数: {len(merged)}")

    # 按公司统计
    if "公司名" in merged.columns:
        print(f"\n公司统计:")
        company_stats = merged.groupby("公司名").size()
        for company, count in company_stats.items():
            print(f"   {company}: {count} 行")

    print(f"\n   列名: {list(merged.columns)}")
    print(f"   输出: {output_path}")

else:
    print("\n❌ 没有成功读取任何数据")
