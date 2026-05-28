"""
tools/excel_tool.py — Excel 文件操作工具（统一版）

单工具多 action 设计，LLM 只需记住一个工具名 `excel`。
所有操作通过 action 参数区分。

功能合并：
 - 原 ExcelOpTool（单格操作）
 - 原 ExcelBatchWriteTool（批量写入）

依赖: openpyxl
"""

import json
import logging
import os
from typing import Any, Dict, List, Optional

from engine.tool.base import BaseTool, ToolParameter, ToolResult

logger = logging.getLogger(__name__)

# ── 全局状态 ──
_open_workbooks: Dict[str, Any] = {}
_file_aliases: Dict[str, str] = {}


def _get_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError("需要 openpyxl，请运行: pip install openpyxl")


def _resolve_path(path_or_alias: str) -> str:
    """解析别名 → 真实路径"""
    if path_or_alias in _file_aliases:
        return _file_aliases[path_or_alias]
    return path_or_alias


class ExcelTool(BaseTool):
    """
    Excel 操作工具（统一版）

    所有操作通过 action 参数区分，一个工具覆盖全部 Excel 场景。

    操作列表：
      connect — 打开文件（需 file_path，可选 alias）
      list_sheets — 列出所有工作表
      read_sheet — 读取工作表（支持分页、智能表头检测）
      get_sheet_info — 获取表头、列数、行数
      find_column — 按列名查找列索引
      write_cell — 写入单个单元格
      write_row — 写入一行
      write_batch — 批量写入多行（原 ExcelBatchWriteTool 功能）
      write_dict — 按字典写入（自动匹配表头）
      save — 保存文件
      close — 关闭文件
    """

    @property
    def name(self) -> str:
        return "excel"

    @property
    def description(self) -> str:
        return (
            "Excel 文件操作工具（基于 openpyxl）。\n"
            "⚠️ 每次调用必须提供 action 参数！\n"
            "\n"
            "📌 标准流程: connect → read_sheet → [分析/修改] → save → close\n"
            "📌 写入流程: connect → read_sheet(了解结构) → write_cell/write_row/write_batch → save\n"
            "\n"
            "操作列表：\n"
            "  connect — 打开文件（需 file_path，可选 alias 设置别名）\n"
            "  list_sheets — 列出所有工作表名称和行列数\n"
            "  read_sheet — 读取工作表（支持分页、智能跳过空列）\n"
            "  get_sheet_info — 获取表头详细信息和数据类型\n"
            "  find_column — 按列名模糊查找列索引\n"
            "  write_cell — 写入单个单元格\n"
            "  write_row — 写入一行（按列号）\n"
            "  write_batch — 批量写入多行（按列号，适合大量数据迁移）\n"
            "  write_dict — 按字典写入（自动匹配表头名称，适合按字段名写入）\n"
            "  save — 保存文件\n"
            "  close — 关闭文件\n"
            "\n"
            "📌 路径别名：connect 时设置 alias，后续操作用别名代替完整路径。\n"
            "\n"
            "📖 使用示例：\n"
            "  # 读 Excel：\n"
            "  1. excel(action='connect', file_path='data.xlsx', alias='x')\n"
            "  2. excel(action='list_sheets', file_path='x')\n"
            "  3. excel(action='read_sheet', file_path='x', sheet_name='Sheet1', header_row=1)\n"
            "  4. excel(action='close', file_path='x')\n"
            "\n"
            "  # 写 Excel（保留格式）：\n"
            "  1. excel(action='connect', file_path='output.xlsx', alias='x')\n"
            "  2. excel(action='write_cell', file_path='x', sheet_name='Sheet1', row=2, column=1, value='Hello')\n"
            "  3. excel(action='write_row', file_path='x', sheet_name='Sheet1', row=3, columns='[1,2,3]', values='[\"a\",\"b\",\"c\"]')\n"
            "  4. excel(action='save', file_path='x')\n"
            "  5. excel(action='close', file_path='x')\n"
            "\n"
            "  # 批量写入（大量数据，省轮次）：\n"
            "  excel(action='write_batch', file_path='x', data='[{\"row\":2,\"columns\":{1:\"A1\",2:\"B1\"}}]')\n"
            "\n"
            "  # 按字典写入（自动匹配表头列名）：\n"
            "  1. read_sheet 先看表头列名\n"
            "  2. excel(action='write_dict', file_path='x', row_data='{\"row\":5,\"data\":{\"姓名\":\"张三\",\"金额\":100}}')\n"
            "  💡 如果 save/close 前忘记写数据，数据不会保存！"
        )

    @property
    def parameters(self) -> List[ToolParameter]:
        return [
            ToolParameter("action", "string",
                          description="⚠️ 必需！操作类型",
                          required=True,
                          enum=["connect", "list_sheets", "read_sheet", "get_sheet_info",
                                "find_column", "write_cell", "write_row", "write_batch",
                                "write_dict", "save", "close"]),
            ToolParameter("file_path", "string",
                          description="文件路径或别名。connect 时传真实路径，后续可用别名",
                          required=False),
            ToolParameter("alias", "string",
                          description="文件别名（action=connect 时设置）",
                          required=False),
            ToolParameter("sheet_name", "string",
                          description="工作表名称",
                          required=False),
            ToolParameter("start_row", "number",
                          description="起始行号（从1开始，read_sheet 用，默认1）",
                          required=False, default=1),
            ToolParameter("max_rows", "number",
                          description="最大返回行数（read_sheet 用，默认 20，设 0 返回全部）",
                          required=False, default=20),
            ToolParameter("header_row", "number",
                          description="表头行号（read_sheet 用，默认1，设0表示无表头）",
                          required=False, default=1),
            ToolParameter("column_name", "string",
                          description="要查找的列名（find_column 用）",
                          required=False),
            ToolParameter("row", "number",
                          description="行号（从1开始，write_cell/write_row 用）",
                          required=False),
            ToolParameter("column", "number",
                          description="列号（从1开始，write_cell 用）",
                          required=False),
            ToolParameter("columns", "string",
                          description="列号数组 JSON（write_row 用，如 '[1,3,5]'）",
                          required=False),
            ToolParameter("value", "string",
                          description="要写入的值（write_cell 用）",
                          required=False),
            ToolParameter("values", "string",
                          description="值数组 JSON（write_row 用，如 '[\"a\",\"b\",\"c\"]'）",
                          required=False),
            ToolParameter("data", "string",
                          description=(
                              "批量写入数据 JSON（write_batch 用）。\n"
                              "格式: '[{\"row\": 2, \"columns\": {1: \"值\", 3: \"值\"}}, ...]'\n"
                              "每项指定行号和列号→值的映射"
                          ),
                          required=False),
            ToolParameter("row_data", "string",
                          description=(
                              "字典格式数据 JSON（write_dict 用）。\n"
                              "格式: '{\"row\": 5, \"data\": {\"姓名\": \"张三\", \"金额\": 100}}'\n"
                              "自动根据表头名称匹配列号"
                          ),
                          required=False),
        ]

    # ═══════════════════════════════════════
    # 主入口
    # ═══════════════════════════════════════

    async def execute(self, call_id: str, **kwargs) -> ToolResult:
        openpyxl = _get_openpyxl()

        if not kwargs:
            return ToolResult.error(call_id, self.name,
                                    self._help_message())

        action = kwargs.get("action", "")

        # 自动推断 action
        if not action:
            action = self._infer_action(kwargs)

        # 解析别名
        raw_path = kwargs.get("file_path", "")
        resolved = _resolve_path(raw_path) if raw_path else raw_path
        if resolved != raw_path:
            logger.info(f"[excel] 别名解析: {raw_path} → {resolved}")
        kwargs["file_path"] = resolved

        # 需要文件的操作：检查是否已连接
        if action in self._actions_need_file() and not resolved:
            return ToolResult.error(call_id, self.name,
                                    self._connect_required_message())

        if action in self._actions_need_file() and resolved and resolved not in _open_workbooks:
            return ToolResult.error(call_id, self.name,
                                    f"❌ 文件未打开: {resolved}\n"
                                    f"📌 请先执行 connect\n"
                                    f"💡 已连接的别名: {list(_file_aliases.keys())}")

        try:
            handler = getattr(self, f"_action_{action}", None)
            if handler is None:
                return ToolResult.error(call_id, self.name,
                                        f"未知 action: {action}。可用: connect, list_sheets, read_sheet, "
                                        f"get_sheet_info, find_column, write_cell, write_row, "
                                        f"write_batch, write_dict, save, close")
            return await handler(call_id, openpyxl, **kwargs)
        except Exception as e:
            logger.exception(f"[excel] {action} 失败")
            return ToolResult.error(call_id, self.name, f"操作失败: {e}")

    # ═══════════════════════════════════════
    # connect — 打开文件
    # ═══════════════════════════════════════

    async def _action_connect(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = kwargs.get("file_path", "")
        alias = kwargs.get("alias", "")

        if not path or not os.path.exists(path):
            return ToolResult.error(call_id, self.name, f"文件不存在: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        _open_workbooks[path] = wb

        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
            })

        alias_name = alias or os.path.basename(path)
        _file_aliases[alias_name] = path
        if alias_name != path:
            _open_workbooks[alias_name] = wb

        return ToolResult.success(call_id, self.name, {
            "path": path,
            "alias": alias_name,
            "sheets": sheets,
            "sheet_count": len(sheets),
            "hint": f"已连接。后续操作可用 file_path='{alias_name}' 代替完整路径",
        })

    # ═══════════════════════════════════════
    # list_sheets — 列出工作表
    # ═══════════════════════════════════════

    async def _action_list_sheets(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column})

        return ToolResult.success(call_id, self.name, {"sheets": sheets, "count": len(sheets)})

    # ═══════════════════════════════════════
    # read_sheet — 读取工作表
    # ═══════════════════════════════════════

    async def _action_read_sheet(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")
        start_row = int(kwargs.get("start_row", 1))
        max_rows = int(kwargs.get("max_rows", 20))
        header_row = int(kwargs.get("header_row", 1))

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name,
                                    f"Sheet 不存在: '{sheet_name}'，可用: {wb.sheetnames}")

        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        # 智能检测有效列
        valid_cols, valid_headers = self._detect_valid_columns(ws, header_row, max_col, max_row)

        # 分页读取
        data_start = max(start_row, header_row + 1 if header_row > 0 else 1)
        if max_rows > 0:
            data_end = min(max_row, data_start + max_rows - 1)
            has_more = data_end < max_row
        else:
            data_end = max_row
            has_more = False

        rows = []
        for r in range(data_start, data_end + 1):
            row_data = {}
            for idx, c in enumerate(valid_cols):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    header_name = valid_headers[idx] if idx < len(valid_headers) else f"列{c}"
                    row_data[header_name] = val
            if row_data:
                rows.append(row_data)

        result = {
            "sheet": sheet_name,
            "total_rows": max_row,
            "total_cols": len(valid_cols),
            "headers": valid_headers,
            "rows": rows,
            "row_count": len(rows),
        }

        if has_more:
            result["hint"] = (
                f"还有 {max_row - data_end} 行未显示。"
                f"用 start_row={data_end + 1} 继续翻页，或用 max_rows=0 读取全部"
            )

        return ToolResult.success(call_id, self.name, result)

    # ═══════════════════════════════════════
    # get_sheet_info — 表头详情
    # ═══════════════════════════════════════

    async def _action_get_sheet_info(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        ws = wb[sheet_name]
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            sample = ws.cell(row=2, column=c).value
            sample_type = type(sample).__name__ if sample is not None else "empty"
            headers.append({
                "col_index": c,
                "col_letter": openpyxl.utils.get_column_letter(c),
                "name": str(val).strip() if val else f"列{c}",
                "sample_type": sample_type,
            })

        return ToolResult.success(call_id, self.name, {
            "sheet": sheet_name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "headers": headers,
        })

    # ═══════════════════════════════════════
    # find_column — 查找列
    # ═══════════════════════════════════════

    async def _action_find_column(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")
        column_name = kwargs.get("column_name", "")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        ws = wb[sheet_name]
        found = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val and column_name.lower() in str(val).lower():
                found.append({
                    "col_index": c,
                    "col_letter": openpyxl.utils.get_column_letter(c),
                    "name": str(val).strip(),
                })

        if found:
            return ToolResult.success(call_id, self.name, {
                "keyword": column_name,
                "matches": found,
                "best_match": found[0],
            })
        else:
            all_headers = [str(ws.cell(row=1, column=c).value or f"列{c}")
                           for c in range(1, ws.max_column + 1)]
            return ToolResult.error(call_id, self.name,
                                    f"未找到匹配 '{column_name}' 的列。可用列: {all_headers}")

    # ═══════════════════════════════════════
    # write_cell — 写入单格
    # ═══════════════════════════════════════

    async def _action_write_cell(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")
        row = int(kwargs.get("row", 0))
        column = int(kwargs.get("column", 0))
        value = kwargs.get("value", "")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")
        if not row or not column:
            return ToolResult.error(call_id, self.name, "需要 row 和 column 参数")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        ws = wb[sheet_name]
        ws.cell(row=row, column=column, value=value)

        return ToolResult.success(call_id, self.name, {
            "cell": f"{openpyxl.utils.get_column_letter(column)}{row}",
            "value": str(value)[:100],
            "status": "已写入，请用 save 保存",
        })

    # ═══════════════════════════════════════
    # write_row — 写入一行
    # ═══════════════════════════════════════

    async def _action_write_row(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")
        row = int(kwargs.get("row", 0))

        try:
            columns = json.loads(kwargs.get("columns", "[]"))
            values = json.loads(kwargs.get("values", "[]"))
        except json.JSONDecodeError:
            return ToolResult.error(call_id, self.name, "columns/values 必须是合法的 JSON 数组")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")
        if not row:
            return ToolResult.error(call_id, self.name, "需要 row 参数")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        if len(columns) != len(values):
            return ToolResult.error(call_id, self.name,
                                    f"列数({len(columns)})与值数({len(values)})不匹配")

        ws = wb[sheet_name]
        written = []
        for col, val in zip(columns, values):
            ws.cell(row=row, column=int(col), value=val)
            written.append(f"{openpyxl.utils.get_column_letter(int(col))}{row}={str(val)[:50]}")

        return ToolResult.success(call_id, self.name, {
            "row": row,
            "cells_written": len(written),
            "preview": written[:10],
            "status": "已写入，请用 save 保存",
        })

    # ═══════════════════════════════════════
    # write_batch — 批量写入（原 ExcelBatchWriteTool）
    # ═══════════════════════════════════════

    async def _action_write_batch(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")

        try:
            data = json.loads(kwargs.get("data", "[]"))
        except json.JSONDecodeError:
            return ToolResult.error(call_id, self.name,
                                    "data 参数必须是合法的 JSON 数组。\n"
                                    "格式: [{\"row\": 2, \"columns\": {1: \"值\", 3: \"值\"}}, ...]")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")
        if not sheet_name:
            return ToolResult.error(call_id, self.name, "需要 sheet_name 参数")
        if not data:
            return ToolResult.error(call_id, self.name, "data 为空，没有要写入的数据")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        ws = wb[sheet_name]
        total_cells = 0
        errors = []

        for row_spec in data:
            row_num = row_spec.get("row")
            columns = row_spec.get("columns", {})
            if not row_num or not columns:
                errors.append(f"跳过无效行定义: {row_spec}")
                continue

            for col_idx, value in columns.items():
                try:
                    ws.cell(row=int(row_num), column=int(col_idx), value=value)
                    total_cells += 1
                except Exception as e:
                    errors.append(f"写入失败 row={row_num}, col={col_idx}: {e}")

        return ToolResult.success(call_id, self.name, {
            "cells_written": total_cells,
            "rows_affected": len(data),
            "errors": errors,
            "error_count": len(errors),
            "status": f"已写入 {total_cells} 个单元格，请用 save 保存",
        })

    # ═══════════════════════════════════════
    # write_dict — 按字段名写入
    # ═══════════════════════════════════════

    async def _action_write_dict(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        sheet_name = kwargs.get("sheet_name", "")

        try:
            row_data = json.loads(kwargs.get("row_data", "{}"))
        except json.JSONDecodeError:
            return ToolResult.error(call_id, self.name,
                                    "row_data 参数必须是合法的 JSON 对象。\n"
                                    "格式: {\"row\": 5, \"data\": {\"姓名\": \"张三\", \"金额\": 100}}")

        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")
        if not sheet_name:
            return ToolResult.error(call_id, self.name, "需要 sheet_name 参数")

        row_num = row_data.get("row")
        data = row_data.get("data", {})
        if not row_num or not data:
            return ToolResult.error(call_id, self.name,
                                    "row_data 格式错误，需要 {\"row\": 行号, \"data\": {\"列名\": 值, ...}}")

        wb = _open_workbooks[path]
        if sheet_name not in wb.sheetnames:
            return ToolResult.error(call_id, self.name, f"Sheet 不存在: '{sheet_name}'")

        ws = wb[sheet_name]

        # 读取表头，建立 列名 → 列号 映射
        header_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                header_map[str(val).strip()] = c

        written = []
        not_found = []
        for field_name, value in data.items():
            if field_name in header_map:
                col = header_map[field_name]
                ws.cell(row=int(row_num), column=col, value=value)
                written.append(f"{field_name}={str(value)[:50]}")
            else:
                # 模糊匹配
                matched = self._fuzzy_match_header(field_name, list(header_map.keys()))
                if matched:
                    col = header_map[matched]
                    ws.cell(row=int(row_num), column=col, value=value)
                    written.append(f"{field_name}→{matched}={str(value)[:50]}")
                else:
                    not_found.append(field_name)

        result = {
            "row": row_num,
            "written": written,
            "written_count": len(written),
        }
        if not_found:
            result["not_found"] = not_found
            result["available_headers"] = list(header_map.keys())
            result["hint"] = f"以下字段在表头中未找到: {not_found}。请检查列名是否正确"

        result["status"] = "已写入，请用 save 保存" if written else "未写入任何数据"
        return ToolResult.success(call_id, self.name, result)

    # ═══════════════════════════════════════
    # save — 保存
    # ═══════════════════════════════════════

    async def _action_save(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        wb.save(path)
        size = os.path.getsize(path) if os.path.exists(path) else 0
        return ToolResult.success(call_id, self.name, {
            "path": path,
            "saved": True,
            "size_kb": round(size / 1024, 1),
        })

    # ═══════════════════════════════════════
    # close — 关闭
    # ═══════════════════════════════════════

    async def _action_close(self, call_id: str, openpyxl, **kwargs) -> ToolResult:
        path = _resolve_path(kwargs.get("file_path", ""))
        if path not in _open_workbooks:
            return ToolResult.error(call_id, self.name, f"文件未打开: {path}")

        wb = _open_workbooks[path]
        wb.close()
        del _open_workbooks[path]

        # 清理别名
        to_remove = [k for k, v in _file_aliases.items() if v == path]
        for k in to_remove:
            del _file_aliases[k]
            _open_workbooks.pop(k, None)

        return ToolResult.success(call_id, self.name, {"path": path, "closed": True})

    # ═══════════════════════════════════════
    # 辅助方法
    # ═══════════════════════════════════════

    @staticmethod
    def _actions_need_file() -> set:
        return {"list_sheets", "read_sheet", "get_sheet_info", "find_column",
                "write_cell", "write_row", "write_batch", "write_dict", "save", "close"}

    @staticmethod
    def _infer_action(kwargs: dict) -> str:
        """自动推断 action（未传 action 时的智能回退）"""
        if kwargs.get("data"):
            return "write_batch"
        if kwargs.get("row_data"):
            return "write_dict"
        if kwargs.get("column_name"):
            return "find_column"
        if kwargs.get("columns") and kwargs.get("values"):
            return "write_row"
        if kwargs.get("row") and kwargs.get("column"):
            return "write_cell"
        if kwargs.get("alias"):
            return "connect"
        if kwargs.get("sheet_name"):
            return "read_sheet"
        if kwargs.get("file_path"):
            return "connect"
        return "read_sheet"

    @staticmethod
    def _detect_valid_columns(ws, header_row: int, max_col: int, max_row: int) -> tuple:
        """智能检测有效列：跳过无意义列名和全空列"""
        valid_cols = []
        valid_headers = []

        for c in range(1, max_col + 1):
            cell_val = ws.cell(row=header_row, column=c).value if header_row > 0 else None
            str_val = str(cell_val).strip() if cell_val else ""

            # 跳过 openpyxl 自动生成的 "列N" 格式
            is_meaningful = str_val and not str_val.startswith("列")

            # 检查该列是否有数据
            has_data = False
            check_start = header_row + 1 if header_row > 0 else 1
            for r in range(check_start, min(check_start + 3, max_row + 1)):
                if ws.cell(row=r, column=c).value is not None:
                    has_data = True
                    break

            if is_meaningful or has_data:
                valid_cols.append(c)
                valid_headers.append(str_val if is_meaningful else f"列{c}")

        # 如果全部被过滤，回退到全部列
        if not valid_cols:
            valid_cols = list(range(1, max_col + 1))
            if header_row > 0:
                valid_headers = [str(ws.cell(row=header_row, column=c).value or f"列{c}")
                                 for c in valid_cols]
            else:
                valid_headers = [f"列{c}" for c in valid_cols]

        return valid_cols, valid_headers

    @staticmethod
    def _fuzzy_match_header(target: str, headers: List[str]) -> Optional[str]:
        """模糊匹配表头（包含/被包含关系）"""
        target_lower = target.lower().strip()
        for h in headers:
            h_lower = h.lower().strip()
            if target_lower in h_lower or h_lower in target_lower:
                return h
        return None

    def _help_message(self) -> str:
        aliases = list(_file_aliases.keys())
        return (
            "❌ 参数为空！请提供 action 参数。\n\n"
            "常用操作:\n"
            "  connect → {\"action\": \"connect\", \"file_path\": \"/path/to/file.xlsx\", \"alias\": \"别名\"}\n"
            "  read_sheet → {\"action\": \"read_sheet\", \"file_path\": \"别名\", \"sheet_name\": \"Sheet1\"}\n"
            "  write_cell → {\"action\": \"write_cell\", \"file_path\": \"别名\", \"sheet_name\": \"Sheet1\", \"row\": 2, \"column\": 1, \"value\": \"内容\"}\n"
            "  write_batch → {\"action\": \"write_batch\", \"file_path\": \"别名\", \"sheet_name\": \"Sheet1\", \"data\": [{\"row\": 2, \"columns\": {1: \"值\"}}]}\n"
            "  write_dict → {\"action\": \"write_dict\", \"file_path\": \"别名\", \"sheet_name\": \"Sheet1\", \"row_data\": {\"row\": 2, \"data\": {\"姓名\": \"张三\"}}}\n"
            "  find_column → {\"action\": \"find_column\", \"file_path\": \"别名\", \"sheet_name\": \"Sheet1\", \"column_name\": \"金额\"}\n"
            "  save → {\"action\": \"save\", \"file_path\": \"别名\"}\n"
            "  close → {\"action\": \"close\", \"file_path\": \"别名\"}\n"
            + (f"\n📌 已连接的文件别名: {aliases}" if aliases else "\n📌 尚未连接任何文件，请先 connect")
        )

    def _connect_required_message(self) -> str:
        aliases = list(_file_aliases.keys())
        return (
            "❌ 尚未连接任何 Excel 文件！\n"
            "📌 请先执行 connect:\n"
            '  {"action": "connect", "file_path": "/path/to/file.xlsx", "alias": "源文件"}\n'
            + (f"💡 已连接的别名: {aliases}" if aliases else "")
        )
